import openpyxl

# Buat nyimpen data dari file masukan.xls
data_masukan = {}

# Buat nyimpen data yang nanti disimpan di luaran.xls
data_luaran = {}


# TODO Baca data dari file masukan.xls
def read_masukan():
    wb = openpyxl.load_workbook("masukan.xlsx")
    sheet = wb['Sheet1']

    for row in sheet.iter_rows(min_row=2, max_row=101, min_col=1, values_only=True):
        id = row[0]
        masukan = {
            "nama_tempat_makan": row[2],
            "rating": row[4],
            "jumlah_menu": row[3],
            "harga_rata_rata": row[6],
        }
        data_masukan[id] = masukan
    wb.close()


# TODO Tulis data hasil olahan observasi ke file luaran.xls
def write_luaran():
    luaran_file = openpyxl.Workbook()
    luaran_sheet = luaran_file.active

    luaran_sheet['A1'] = 'ID'
    luaran_sheet['B1'] = 'Nama Tempat Makan'
    luaran_sheet['C1'] = 'NK'

    row = 1
    for index in data_luaran:
        resto_obj = data_luaran[index]
        row += 1
        luaran_sheet[f'A{row}'] = index
        luaran_sheet[f'B{row}'] = resto_obj['nama_tempat_makan']
        luaran_sheet[f'C{row}'] = resto_obj['NK']

    luaran_file.save('luaran.xlsx')
    luaran_file.close()


# Attribute Grouping:
# Rating:
#   - Enak              : > 3.7
#   - Biasa             : 2 - 4
#   - Kurang Enak       : < 2.3
#
# Rata-rata Harga:
#   - Mahal             : > 50 ribu
#   - Dapat Diterima    : 15 - 55 ribu
#   - Murah       : < 20 ribu
#
# Jumlah Menu:
#   - Sangat Variatif   : > 33 menu
#   - Variatif          : > 5 - 35 menu
#   - Kurang Variatif   : < 7

rating_sets = {
    'Bagus': {
        'upper': 5,
        'lower': 3.7
    },
    'Biasa': {
        'upper': 4,
        'lower': 2
    },
    'Buruk': {
        'upper': 2.3,
        'lower': 0
    }
}

harga_sets = {
    'Mahal': {
        'upper': 100000,
        'lower': 50000
    },
    'Diterima': {
        'upper': 55000,
        'lower': 17000
    },
    'Murah': {
        'upper': 20000,
        'lower': 0
    },

}

label_sets = {
    'Oke': {
        'upper': 1,
        'lower': 4 / 9
    },
    'Tidak': {
        'lower': 0,
        'upper': 5 / 9
    }
}


def get_category_sets(value: float, sets: dict):
    sets_area = []
    for set_type in sets.keys():
        type_obj = sets[set_type]
        upper = type_obj['upper']
        lower = type_obj['lower']

        if lower <= value <= upper:
            sets_area.append(set_type)
    return sets_area


def get_grouping(value: float, groups: list[int]):
    for i in range(0, len(groups)):
        a = groups[i]
        b = groups[i + 1]
        lower = min(a, b)
        upper = max(a, b)

        if lower < value <= upper:
            return lower, upper


def calc_group(value: float, lower: float, upper: float):
    low = -(value - upper) / (upper - lower)
    high = (value - lower) / (upper - lower)
    return low, high


def calc_group_v2(value: float, categories: list, lower: float, upper: float):
    if len(categories) == 1:
        return {categories[0]: 1}
    else:
        low = -(value - upper) / (upper - lower)
        high = (value - lower) / (upper - lower)
        return {categories[0]: high, categories[1]: low}


# TODO fungsi fuzzifikasi disini
# contoh di slide halaman 54
def fuzzification():
    fuzzy_result = {}

    for resto_id in data_masukan.keys():
        resto_obj = data_masukan[resto_id]

        rating = resto_obj['rating']
        harga_rata_rata = resto_obj['harga_rata_rata']

        rating_lower, rating_upper = get_grouping(rating, rating_groups)
        harga_lower, harga_upper = get_grouping(harga_rata_rata, harga_groups)

        rating_fuzzy = calc_group(rating, rating_lower, rating_upper)
        harga_fuzzy = calc_group(harga_rata_rata, harga_lower, harga_upper)

        rating_result = {'lower': rating_fuzzy[0], 'upper': rating_fuzzy[1]}
        harga_result = {'lower': harga_fuzzy[0], 'upper': harga_fuzzy[1]}

        result = {
            'rating': rating_result,
            'harga': harga_result
        }

        fuzzy_result[resto_id] = result

    return fuzzy_result


def attribute_range_filter(categories: list[str], category_sets: dict):
    filtered = {}

    if len(categories) == 1:
        category = categories[0]
        filtered['lower'] = category_sets[category]['lower']
        filtered['upper'] = category_sets[category]['upper']
    elif len(categories) == 2:
        category_1 = categories[0]
        category_2 = categories[1]
        filtered['lower'] = category_sets[category_1]['lower']
        filtered['upper'] = category_sets[category_2]['upper']

    return filtered


def fuzzification_v2():
    fuzzy_result = {}

    for resto_id in data_masukan.keys():
        resto_obj = data_masukan[resto_id]
        rating = resto_obj['rating']
        harga = resto_obj['harga_rata_rata']

        rating_categories = get_category_sets(rating, rating_sets)
        rating_filter = attribute_range_filter(rating_categories, rating_sets)
        rating_result_values = calc_group_v2(rating, rating_categories, rating_filter['lower'], rating_filter['upper'])
        rating_result = {
            'value': rating,
            'range': {
                'category': rating_categories,
                'range': {
                    'lower': rating_filter['lower'],
                    'upper': rating_filter['upper']
                }
            },
            'result': rating_result_values
        }

        harga_categories = get_category_sets(harga, harga_sets)
        harga_filter = attribute_range_filter(harga_categories, harga_sets)
        harga_result_values = calc_group_v2(harga, harga_categories, harga_filter['lower'], harga_filter['upper'])
        harga_result = {
            'value': harga,
            'range': {
                'category': harga_categories,
                'range': {
                    'lower': harga_filter['lower'],
                    'upper': harga_filter['upper']
                }
            },
            'result': harga_result_values
        }

        resto_fuzzy = {
            'rating': rating_result,
            'harga': harga_result
        }

        fuzzy_result[resto_id] = resto_fuzzy

    return fuzzy_result


# TODO fungsi inference disini
# contoh di slide halaman 55-57
def inference(fuzzy_result: dict):
    inferece_result = {}

    for resto_id in fuzzy_result.keys():
        resto_obj = fuzzy_result[resto_id]
        rating = resto_obj['rating']
        harga = resto_obj['harga']

        rating_fuzzy = rating['result']
        harga_fuzzy = harga['result']

        inf_result = []

        for cat_rating in rating_fuzzy:
            for cat_harga in harga_fuzzy:
                cat_rating_value = rating_fuzzy[cat_rating]
                cat_harga_value = harga_fuzzy[cat_harga]

                if cat_rating == 'Bagus' and cat_harga == 'Murah':
                    inf = {'category': 'Oke', 'NK': inference_conjunction(cat_rating_value, cat_harga_value)}
                elif cat_rating == 'Bagus' and cat_harga == 'Diterima':
                    inf = {'category': 'Oke', 'NK': inference_conjunction(cat_rating_value, cat_harga_value)}
                elif cat_rating == 'Bagus' and cat_harga == 'Mahal':
                    inf = {'category': 'Oke', 'NK': inference_conjunction(cat_rating_value, cat_harga_value)}
                elif cat_rating == 'Biasa' and cat_harga == 'Murah':
                    inf = {'category': 'Oke', 'NK': inference_conjunction(cat_rating_value, cat_harga_value)}
                elif cat_rating == 'Biasa' and cat_harga == 'Diterima':
                    inf = {'category': 'Tidak', 'NK': inference_conjunction(cat_rating_value, cat_harga_value)}
                elif cat_rating == 'Biasa' and cat_harga == 'Mahal':
                    inf = {'category': 'Tidak', 'NK': inference_conjunction(cat_rating_value, cat_harga_value)}
                elif cat_rating == 'Buruk' and cat_harga == 'Murah':
                    inf = {'category': 'Tidak', 'NK': inference_conjunction(cat_rating_value, cat_harga_value)}
                elif cat_rating == 'Buruk' and cat_harga == 'Diterima':
                    inf = {'category': 'Tidak', 'NK': inference_conjunction(cat_rating_value, cat_harga_value)}
                elif cat_rating == 'Buruk' and cat_harga == 'Mahal':
                    inf = {'category': 'Tidak', 'NK': inference_conjunction(cat_rating_value, cat_harga_value)}

                inf_result.append(inf)
        inferece_result[resto_id] = inference_disjunction(inf_result)

    return inferece_result


def inference_conjunction(rating_value: float, harga_value: float):
    return min(rating_value, harga_value)


def inference_disjunction(inf_result: list[dict]):
    result: dict = {}

    for i in range(0, len(inf_result)):
        cat_obj = inf_result[i]
        category = cat_obj['category']
        nk = cat_obj['NK']

        if category in result:
            result[category] = max(nk, result[category])
        else:
            result[category] = nk

    return result


# TODO fungsi defuzzifikasi disini
# contoh di slide halaman 58
def defuzzification(inference_result: dict):
    deffuzzi_result = {}
    for resto_id in data_masukan.keys():
        resto_inference: dict = inference_result[resto_id]
        label_categories = list(resto_inference.keys())

        top_section = 0
        bottom_section = 0
        for label_result in resto_inference:
            label_inf_value = resto_inference[label_result]
            label_ranges = attribute_range_filter(label_categories, label_sets)

            ten_area_floor, ten_area_ceil = get_label_ten_area(label_ranges['lower'], label_ranges['upper'])
            area_amount, area_result = calculate_ten_area(ten_area_floor, ten_area_ceil)

            top, bottom = calculate_pre_deffuzi(label_inf_value, area_result, area_amount)

            top_section += top
            bottom_section += bottom

        resto_deffuzi = calculate_deffuzi(top_section, bottom_section)*10
        deffuzzi_result[resto_id] = resto_deffuzi
    return deffuzzi_result


def calculate_deffuzi(top_section: float, bottom_section: float):
    return top_section / bottom_section


def calculate_pre_deffuzi(label_value: float, area_result: float, area_amount: int):
    top_result = area_result * 10 * label_value
    bottom_result = label_value * area_amount
    return top_result, bottom_result


def calculate_ten_area(floor: float, ceil: float):
    increment = 0
    area_amount: int = 0
    area_result = 0
    for i in range(0, 10):
        if floor <= increment <= ceil:
            area_result += increment
        increment += 0.1
        area_amount += 1
    return area_amount, area_result


def get_label_ten_area(lower: float, upper: float):
    result_floor = -1
    result_ceil = -1

    for i in range(0, 10):
        floor = 0.1 * i
        ceil = 0.1 * (i + 1)

        if result_floor == -1:
            if ceil >= lower:
                result_floor = floor
        if result_ceil == -1:
            if ceil >= upper:
                result_ceil = ceil
        if result_floor != -1 and result_ceil != -1:
            break

    return result_floor, result_ceil


def sort_by_nk(deffuzification_result: dict[any, float]):
    return sorted(deffuzification_result, key=deffuzification_result.get, reverse=True)


if __name__ == '__main__':
    read_masukan()
    fuzzi = fuzzification_v2()
    infe = inference(fuzzi)
    deffuzi = defuzzification(infe)

    for i in sort_by_nk(deffuzi):
        resto_obj = data_masukan[i]
        data_luaran[i] = {
            'nama_tempat_makan': resto_obj['nama_tempat_makan'],
            'NK': deffuzi[i]
        }

    write_luaran()
